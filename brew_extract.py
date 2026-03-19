"""
7BREW P&L Data Extraction Engine
Reads all period Excel files and returns a consolidated DataFrame.
"""

import openpyxl
import pandas as pd
import re
import os
import glob
from datetime import datetime

STAND_DATES_FILE = '7Crew_Stand_Dates.xlsx'

def load_stand_dates(folder):
    """
    Load the stand dates reference file.
    Returns:
      - upcoming_ids: set of stand_id strings for stands not yet open
      - opening_dates: dict of stand_id -> opening date string (for future use)
    """
    path = os.path.join(folder, STAND_DATES_FILE)
    if not os.path.exists(path):
        return set(), {}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    upcoming_ids = set()
    opening_dates = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        stand_raw, region, open_date = (row[0], row[1], row[2]) if len(row) >= 3 else (row[0], None, None)
        if not stand_raw:
            continue
        # Extract stand number from format like "Lubbock (#134)" → "000134"
        m = re.search(r'#(\d+)', str(stand_raw))
        if m:
            stand_id = m.group(1).zfill(6)
            upcoming_ids.add(stand_id)
            if open_date:
                opening_dates[stand_id] = str(open_date)

    print(f"  Stand dates file: {len(upcoming_ids)} upcoming stands identified")
    return upcoming_ids, opening_dates

METRIC_ROWS = {
    'Gross Sales':              16,
    'Net Sales':                22,
    'COGS':                     33,
    'Gross Profit':             34,
    'Total Hourly Labor':       41,
    'Total Management':         45,
    'Total Labor':              46,
    'Total Benefits & Tax':     52,
    'Total Labor & Benefits':   53,
    'Gross Margin':             54,
    'Controllable Expense':     68,
    'Total Utilities':          74,
    'Total R&M':                80,
    'Total Fixed Expense':      100,
    'Total Occupancy':          104,
    'Total Marketing':          113,
    'Unit Level EBITDAR':       114,
    'Total Rent':               115,
    'Store Level EBITDA':       116,
    'EBITDA':                   129,
    'Preopening Expense':       130,
    'Net Income':               140,
}

SKIP_STANDS = {
    'All Stands', 'ENT Corp Office', 'Florida Corp Office',
    'All Stands & Offices', 'All Stands & Office'
}

def parse_period_label(filename):
    """Extract period label from filename like 'P1\'25' → ('P1\'25', 1, 2025)"""
    basename = os.path.basename(filename)
    m = re.search(r"P(\d+)'(\d+)", basename)
    if m:
        period_num = int(m.group(1))
        year = 2000 + int(m.group(2))
        label = f"P{period_num}'{m.group(2)}"
        return label, period_num, year
    return None, None, None

def parse_stand_name(raw):
    """Parse '000134 Lubbock, TX - 1' → (id, city, state, number, region)"""
    raw = str(raw).strip()
    m = re.match(r'^(\d+)\s+(.+),\s+([A-Z]{2})\s*-\s*(\d+)$', raw)
    if m:
        stand_id = m.group(1)
        city = m.group(2).strip()
        state = m.group(3)
        num = int(m.group(4))
        fl_states = {'FL'}
        region = 'Florida' if state in fl_states else 'ENT'
        display = f"{city}, {state} #{num}"
        return stand_id, city, state, num, region, display
    return raw, raw, '', 1, 'ENT', raw

def extract_file(filepath):
    """Extract all stand metrics from one period file. Returns list of row dicts."""
    label, period_num, year = parse_period_label(filepath)
    if label is None:
        return []

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Get period end date
    date_cell = list(ws.iter_rows(min_row=4, max_row=4, min_col=2, max_col=2, values_only=True))[0][0]
    try:
        if isinstance(date_cell, datetime):
            period_date = date_cell
        else:
            period_date = datetime.strptime(str(date_cell), '%m/%d/%Y')
    except Exception:
        period_date = None

    # Get stand names and column positions from row 8
    row8 = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
    stand_cols = {}
    for col_idx, val in enumerate(row8, 1):
        if val and str(val).strip() not in ['', ' ']:
            name = str(val).strip()
            if name not in SKIP_STANDS:
                stand_cols[col_idx] = name

    # Read all metric rows at once for efficiency
    all_row_data = {}
    for metric, row_num in METRIC_ROWS.items():
        row_vals = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        all_row_data[metric] = row_vals

    records = []
    for col_idx, stand_raw in stand_cols.items():
        stand_id, city, state, stand_num, region, display = parse_stand_name(stand_raw)
        row = {
            'period_label': label,
            'period_num': period_num,
            'year': year,
            'period_date': period_date,
            'stand_raw': stand_raw,
            'stand_id': stand_id,
            'city': city,
            'state': state,
            'stand_num': stand_num,
            'region': region,
            'display_name': display,
        }
        for metric, row_vals in all_row_data.items():
            val = row_vals[col_idx - 1]
            row[metric] = float(val) if isinstance(val, (int, float)) else 0.0
        records.append(row)

    return records

def extract_all(input_folder, file_pattern='7BREW Income Statement Side By Side PTD All*.xlsx'):
    """Extract all period files from folder, return consolidated DataFrame."""
    pattern = os.path.join(input_folder, file_pattern)
    files = sorted(glob.glob(pattern))
    if not files:
        raise FileNotFoundError(f"No files found matching: {pattern}")

    all_records = []
    for f in files:
        records = extract_file(f)
        all_records.extend(records)
        label, _, _ = parse_period_label(f)
        print(f"  Loaded {label}: {len(records)} stands")

    df = pd.DataFrame(all_records)

    # Sort chronologically
    df['sort_key'] = df['year'] * 100 + df['period_num']
    df = df.sort_values(['sort_key', 'stand_id']).reset_index(drop=True)

    return df

def assign_cohorts(df):
    """
    Assign each stand to an opening cohort based on its first appearance in the data.
    Also compute 'periods_open' (how many periods since first appearance).
    """
    min_sk = df['sort_key'].min()

    # Find first sort_key for each stand
    first_sk = df.groupby('stand_id')['sort_key'].min()

    def period_to_quarter(pn):
        if pn <= 3:   return 1
        if pn <= 6:   return 2
        if pn <= 9:   return 3
        return 4

    def cohort_label(sk):
        if sk == min_sk:
            return "Legacy (Pre-Data)"
        yr = sk // 100
        pn = sk % 100
        q = period_to_quarter(pn)
        return f"Q{q}'{str(yr)[-2:]}"

    cohort_map = {sid: cohort_label(sk) for sid, sk in first_sk.items()}
    df['cohort'] = df['stand_id'].map(cohort_map)
    df['first_sort_key'] = df['stand_id'].map(first_sk)

    # Periods open = rank of current period for this stand (1 = first period open)
    df['periods_open'] = df.groupby('stand_id')['sort_key'].rank(method='dense').astype(int)

    # Ramp flag: open fewer than 4 periods
    df['is_ramp'] = df['periods_open'] <= 4

    return df

def add_derived_metrics(df):
    """Add % of Net Sales metrics and other derived fields."""
    ns = df['Net Sales'].replace(0, float('nan'))

    df['COGS %'] = df['COGS'] / ns
    df['Labor %'] = df['Total Labor & Benefits'] / ns
    df['Hourly Labor %'] = df['Total Labor'] / ns
    df['Gross Margin %'] = df['Gross Margin'] / ns
    df['Store EBITDA %'] = df['Store Level EBITDA'] / ns
    df['EBITDA %'] = df['EBITDA'] / ns
    df['Controllable %'] = df['Controllable Expense'] / ns
    df['Rent %'] = df['Total Rent'] / ns
    df['Marketing %'] = df['Total Marketing'] / ns

    return df

def build_dataset(input_folder, file_pattern='7BREW Income Statement Side By Side PTD All*.xlsx'):
    """Full pipeline: extract → cohorts → derived metrics → return DataFrame."""
    print(f"Scanning: {input_folder}")
    df = extract_all(input_folder, file_pattern)

    # Load stand dates reference file (source of record for upcoming/future stands)
    upcoming_ids, opening_dates = load_stand_dates(input_folder)

    # Exclude upcoming stands by stand_id from the reference file
    if upcoming_ids:
        before = len(df)
        df = df[~df['stand_id'].isin(upcoming_ids)].reset_index(drop=True)
        print(f"  Excluded {before - len(df)} records for {len(upcoming_ids)} upcoming stands")

    # Also drop any remaining stand-period with Net Sales <= $50K (safety net)
    before = len(df)
    df = df[df['Net Sales'] > 50_000].reset_index(drop=True)
    if before > len(df):
        print(f"  Filtered out {before - len(df)} additional records with Net Sales ≤ $50K")

    # Store opening dates on the dataframe for future use
    if opening_dates:
        df['scheduled_open_date'] = df['stand_id'].map(opening_dates)

    df = assign_cohorts(df)
    df = add_derived_metrics(df)
    print(f"Total records: {len(df)} ({df['stand_id'].nunique()} unique stands, {df['period_label'].nunique()} periods)")
    return df
